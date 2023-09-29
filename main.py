import openpyxl 
from openpyxl import Workbook
import os 
import pandas as pd
import psycopg2
import psycopg2.extras


# access to data in Excel:
dir_path = os.path.dirname(os.path.realpath(__file__))
wb = openpyxl.load_workbook(dir_path + r"\\energy_data.xlsx", "r")

ws_hydro = wb['Hydro Generation - TWh']
ws_ren = wb['Renewable power - TWh']
ws_nuc = wb['Nuclear Generation - TWh']

ws_oil = wb['Elec Gen from Oil']
ws_gas = wb['Elec Gen from Gas']
ws_coal = wb['Elec Gen from Coal']
ws_other = wb['Elec Gen from Other'] 

ws_total = wb['Electricity Generation']

class EnergyData:
    def __init__(self, country):
        self.country = country

    def valid_country(self):
        for row in range(5, 110): 
            cell_in_column_A = ws_hydro.cell(row=row, column=1).value
            if cell_in_column_A == self.country:
                return True 
        return False  

#methods to collect generation mix in TWh from multiple spreadsheets

    def fetch_data65(self, worksheet, start_year, end_year):
        values65 = []

        start_col = start_year - 1965 + 2
        end_col = end_year - 1965 + 2

        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1)
            if cell_in_column_A.value == self.country:
                row_values = [cell.value for cell in row]
                values65.append(row_values)

        return values65

    def fetch_data85(self,  worksheet, start_year, end_year):
        values85 = []

        start_col = start_year - 1985 + 2
        end_col = end_year - 1985 + 2

        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1)
            if cell_in_column_A.value == self.country:
                row_values = [cell.value for cell in row]
                values85.append(row_values)

        return values85

    def histdata(self, start_year, end_year):
        hy_values = self.fetch_data65(ws_hydro, start_year, end_year)[0]
        ren_values = self.fetch_data65(ws_ren, start_year, end_year)[0]
        nuc_values = self.fetch_data65(ws_nuc, start_year, end_year)[0]
        oil_values = self.fetch_data85(ws_oil, start_year, end_year)[0]
        gas_values = self.fetch_data85(ws_gas, start_year, end_year)[0]
        coal_values = self.fetch_data85(ws_coal, start_year, end_year)[0]
        other_values = self.fetch_data85(ws_other, start_year, end_year)[0]
        
        date = list(range(start_year, end_year+1))
        fuelmix = ["Hydro", "Renewables", "Nuclear", "Oil", "Gas", "Coal", "Other"]


        table = [[" "] + fuelmix]
        for i, year in enumerate(date):
            row_values = [year] + [hy_values[i], ren_values[i], nuc_values[i], oil_values[i], gas_values[i], coal_values[i], other_values[i]]
            table.append(row_values)

        df = pd.DataFrame(table[1:], columns=table[0]).set_index(' ')
        df.reset_index(drop=True)
        return df.T

# method to get generation mix percentage for one country 

    def get_the_share(self, country, year):
        table = self.histdata(year, year)
        total_sum = table[year].sum()
        table['Share, %'] = ((table[year]/ total_sum) * 100).round(1)
        table = table.sort_values(by='Share, %', ascending = False)
        return table

# method to get generation mix percentage for multiple countries 

    def compare_countries(self):
        try:
            year = int(input("Insert the year:\n"))
            if year < 1985 or year > 2022:
                raise ValueError("The year must be between 1985 and 2022.")
            
            country_list = input("Insert several countries separated by commas:\n").split(',')
            country_list = [country.strip() for country in country_list]
            
            all_countries_valid = all(EnergyData(country).valid_country() for country in country_list)
            
            if all_countries_valid == False:
                print("One or more countries in the list are invalid.")
                return

            table = {}

            for country in country_list:
                a = EnergyData(country)
                table[country] = a.get_the_share(country, year).rename(columns={"Share, %": country}).drop([year], axis=1)

            result = table[country_list[0]]

            for country in country_list[1:]:
                result = pd.merge(result, table[country], left_index=True, right_index=True)

            print(result)
        except ValueError as e:
            print(e)

# method to export hist_data to excel

    def export_to_excel(self, start_year, end_year, country):
        data = self.histdata(start_year, end_year)
        df = pd.DataFrame(data)
        name = f"{country}.xlsx"
        file_path = os.path.join(dir_path, name)
        df.to_excel(file_path, index=True)
        print(f"The data is exported to the '{country}.xlsx' file.")

#methods to work with SQL Postgress

    def connect_to_postgresql(self):
        DB_NAME = "energy_mix"
        USER = "postgres"
        PASSWORD = "root"
        HOST = "localhost"
        PORT = "5432" 

        try:
            global connection
            connection = psycopg2.connect(
                dbname = DB_NAME, 
                user = USER,
                password = PASSWORD,
                host = HOST,
                port = PORT
            )
        except Exception as e:
            print(f"Error: {e}")
        
        global cursor
        cursor = connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    def create_a_table(self, table_name: str, start_year, end_year):
        try:
            self.connect_to_postgresql()
            date = list(range(start_year, end_year+1))
            query = f'''
                create table if not exists {table_name} (
                    fuel_type varchar(20) primary key,
                    {', '.join([f'"{year}" numeric' for year in date])}
                );
                '''
            cursor.execute(query) 
            connection.commit()
            return f"The table {table_name} is created."
        except Exception as e:
            None

    def add_data_to_a_table(self, table_name: str, start_year, end_year):
        try:
            self.connect_to_postgresql()

            date = list(range(start_year, end_year+1))
            data = self.histdata(start_year, end_year)
            all_values = []

            for fuel_type, row in data.iterrows():
                values = list(row)
                all_values.append((fuel_type, values))

            query = f'''
                insert into {table_name} (fuel_type, {', '.join([f'"{year}"' for year in date])})
                values
                {', '.join([f"('{fuel_type}', {', '.join([str(v) for v in values])})" for fuel_type, values in all_values])};
            '''

            cursor.execute(query) 
            connection.commit()
            return f"The data is added into '{table_name}' table."
        except Exception as e:
            return f"Data export failed: the table '{table_name}' doesn't exist."
    
    def delete_the_table(self, table_name: str):
        try:
            self.connect_to_postgresql()
            query = f'''
                drop table {table_name};
            '''

            cursor.execute(query) 
            connection.commit()
            return f"The table '{table_name}' is deleted."
        except Exception as e:
            return f"Table deletion failed: the table '{table_name}' was not found."


 





















